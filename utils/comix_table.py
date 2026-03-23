import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl.styles import Font
from openpyxl import load_workbook

from utils.spreadsheet import write_title, autofit_columns, hide_cols

def make_chapters_df(items: list) -> pd.DataFrame:
    df = pd.DataFrame([{
        "chapter_id": item["chapter_id"],
        "number": item["number"],
        "name": item["name"],
        "created_at": datetime.fromtimestamp(item["created_at"]).strftime("%Y-%m-%d"),
        "scanlation_group": item["scanlation_group"]["name"] if item["scanlation_group"] else "Unknown"
    } for item in items])
    return df.iloc[::-1].reset_index(drop=True)

def make_spreedsheet(df: pd.DataFrame, title: str, path: Path):

    with pd.ExcelWriter(path, engine="openpyxl") as writer:

        # Write DataFrame to Excel
        df.to_excel(writer, index=False)

        # Load excel workbook
        wb = writer.book
        ws = writer.sheets["Sheet1"]

        # Hide unwanted columns and autofit
        hide_cols(ws, ["chapter_id"])
        autofit_columns(ws)

        # Add title
        write_title(ws, title, cols = 5)

        # Bold headers
        for cell in ws[2]:
            cell.font = Font(bold=True)

        wb.save(path)

def read_spreadsheet(path: Path) -> tuple[str, pd.DataFrame]:
    
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    title = ws["A1"].value

    df = pd.read_excel(path, header=1, usecols=["chapter_id", "number", "name"])

    # change "nan" into None in the "name" field
    df["name"] = df["name"].where(pd.notna(df["name"]), None)

    return title, df
