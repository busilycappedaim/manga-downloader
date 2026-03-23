from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import Alignment

# Helper functions for Excel formatting
def autofit_columns(ws: Worksheet) -> None:
    for col in ws.columns:
        col_letter = col[0].column_letter
        if ws.column_dimensions[col_letter].hidden:
            continue

        max_length = max(
            len(str(cell.value)) if cell.value else 0
            for cell in col
        )
        ws.column_dimensions[col_letter].width = max_length + 2

# Helper functions for Excel formatting
def hide_cols(ws: Worksheet, hidden_cols: list[str]) -> None:

    header_map = {
        cell.value: cell.column_letter
        for cell in ws[1]
        if cell.value is not None
    }

    for col_name in hidden_cols:
        col_letter = header_map.get(col_name)
        if col_letter:
            ws.column_dimensions[col_letter].hidden = True

def write_title(ws: Worksheet, title: str, cols: int) -> None:
    ws.insert_rows(1)
    ws.merge_cells(f"A1:{get_column_letter(cols)}1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")